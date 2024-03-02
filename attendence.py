import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import openpyxl

class AttendanceTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Attendance Tracker")
        self.root.geometry("800x500")

        self.attendance_records = {}
        self.subjects = []
        self.load_previous_attendance()

        # Create a frame for input widgets
        input_frame = tk.Frame(self.root)
        input_frame.pack(pady=10)

        # Student Name Entry
        tk.Label(input_frame, text="Enter Student Name:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.student_name = tk.Entry(input_frame)
        self.student_name.grid(row=0, column=1, padx=5, pady=5)

        # Subject Code Entry
        tk.Label(input_frame, text="Enter Subject Code:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.subject_code = tk.Entry(input_frame)
        self.subject_code.grid(row=1, column=1, padx=5, pady=5)

        # Mark Attendance Buttons
        mark_frame = tk.Frame(self.root)
        mark_frame.pack(pady=10)
        tk.Button(mark_frame, text="Mark Present", command=lambda: self.mark_attendance("P")).grid(row=0, column=0, padx=5, pady=5)
        tk.Button(mark_frame, text="Mark Absent", command=lambda: self.mark_attendance("A")).grid(row=0, column=1, padx=5, pady=5)

        # Add Subject Button
        add_subject_frame = tk.Frame(self.root)
        add_subject_frame.pack(pady=10)
        tk.Button(add_subject_frame, text="Add Subject", command=self.add_subject).pack(padx=5, pady=5)

        # Attendance Table
        table_frame = tk.Frame(self.root)
        table_frame.pack(pady=10)
        self.attendance_table = ttk.Treeview(table_frame, columns=["Subject"] + self.subjects, show="headings")
        self.attendance_table.pack(side="left")

        # Set up treeview columns
        self.attendance_table.heading("#0", text="Student Name")
        self.attendance_table.heading("Subject", text="Subject")
        for subject in self.subjects:
            self.attendance_table.heading(subject, text=subject)

        # Add a scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.attendance_table.yview)
        scrollbar.pack(side="right", fill="y")
        self.attendance_table.configure(yscrollcommand=scrollbar.set)

        # Get Attendance Percentage Button
        tk.Button(self.root, text="Get Attendance Percentage", command=self.get_attendance_percentage).pack(pady=10)

    def mark_attendance(self, status):
        student_name = self.student_name.get()
        subject_code = self.subject_code.get()

        if not student_name or not subject_code:
            messagebox.showerror("Error", "Please enter student name and subject code")
            return

        if student_name not in self.attendance_records:
            self.attendance_records[student_name] = {}

        if subject_code not in self.attendance_records[student_name]:
            self.attendance_records[student_name][subject_code] = []

        self.attendance_records[student_name][subject_code].append(status)

        if subject_code not in self.subjects:
            self.subjects.append(subject_code)
            # Update table headings
            self.attendance_table["columns"] = ["Subject"] + self.subjects
            self.attendance_table.heading(subject_code, text=subject_code)

        self.update_table()
        self.student_name.delete(0, tk.END)
        self.subject_code.delete(0, tk.END)

        self.save_attendance_to_excel()

    def add_subject(self):
        subject_code = self.subject_code.get()

        if not subject_code:
            messagebox.showerror("Error", "Please enter a subject code")
            return

        if subject_code not in self.subjects:
            self.subjects.append(subject_code)
            # Update table headings
            self.attendance_table["columns"] = ["Subject"] + self.subjects
            self.attendance_table.heading(subject_code, text=subject_code)

        self.subject_code.delete(0, tk.END)

    def update_table(self):
        # Clear previous entries
        for row in self.attendance_table.get_children():
            self.attendance_table.delete(row)
        
        # Update table with attendance records
        for student_name, records in self.attendance_records.items():
            for subject in records.keys():
                row_data = [student_name, subject]
                for sub in self.subjects:
                    if sub == subject:
                        row_data.append(records[subject].count("P") / len(records[subject]) * 100)
                    else:
                        row_data.append("-")
                self.attendance_table.insert("", "end", values=row_data)

    def get_attendance_percentage(self):
        student_name = self.student_name.get()

        if not student_name:
            messagebox.showerror("Error", "Please enter a student name")
            return

        if student_name not in self.attendance_records:
            messagebox.showerror("Error", "No attendance record found for this student")
            return

        for subject, records in self.attendance_records[student_name].items():
            total_attendances = len(records)
            present_count = records.count("P")
            if total_attendances == 0:
                attendance_percentage = 0
            else:
                attendance_percentage = (present_count / total_attendances) * 100
            messagebox.showinfo(f"{subject} Attendance Percentage", f"{student_name}'s attendance percentage in {subject} is {attendance_percentage:.2f}%")

    def load_previous_attendance(self):
        try:
            wb = openpyxl.load_workbook("attendance.xlsx")
            sheet = wb.active

            # Load attendance records from Excel
            for row in sheet.iter_rows(min_row=2, values_only=True):
                student_name, subject, status = row
                if student_name not in self.attendance_records:
                    self.attendance_records[student_name] = {}
                if subject not in self.attendance_records[student_name]:
                    self.attendance_records[student_name][subject] = []
                self.attendance_records[student_name][subject].append(status)

            wb.close()
        except FileNotFoundError:
            messagebox.showwarning("Warning", "Attendance file not found. Creating a new one.")

    def save_attendance_to_excel(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Student Name", "Subject", "Status"])

        for student_name, records in self.attendance_records.items():
            for subject, statuses in records.items():
                for status in statuses:
                    sheet.append([student_name, subject, status])

        wb.save("attendance.xlsx")
        messagebox.showinfo("Success", "Attendance saved to Excel file.")
        
if __name__ == "__main__":
    root = tk.Tk()
    AttendanceTracker(root)
    root.mainloop()
